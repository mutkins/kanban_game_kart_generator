from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Side

THICK_BORDER = Side(style='thick', color='000000')
THIN_BORDER = Side(style='thin', color='000000')

MED_GREEN_FILL = PatternFill("lightGray", start_color="92D050")
MED_BLUE_FILL = PatternFill("lightGray", start_color="00B0F0")
MED_YELLOW_FILL = PatternFill("lightGray", start_color="FFFF00")
MED_RED_FILL = PatternFill("lightGray", start_color="FF0000")
MED_BLACK_FILL = PatternFill("lightGray", start_color="000000")
NO_FILL = PatternFill("lightGray", start_color="FFFFFF")

GREEN_FILL = PatternFill("darkGray", start_color="92D050")
BLUE_FILL = PatternFill("darkGray", start_color="00B0F0")
RED_FILL = PatternFill("darkGray", start_color="FF0000")
