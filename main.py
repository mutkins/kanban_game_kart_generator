from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
import modificators_catalog
from Modificators import Modificators
from Styles import MED_BIRUZA
from Troubles import Troubles
from tools import *
from Stories import Stories
workbook = Workbook()

sheet = workbook.active



stories = Stories()

stories.generate_u_stories(letter_index='S', count=50)
stories.generate_f_stories(letter_index='F', count=8)
stories.generate_o_stories(letter_index='O', count=16)
stories.generate_e_stories(letter_index='E', count=14)

for stories_list in stories.stories_list:
    sheet = workbook.create_sheet(title=stories_list[0].letter_index)
    place_story_on_sheet(stories_list=stories_list, sheet=sheet)

place_card_back(sheet=workbook.create_sheet(title='S_back'), color=MED_GREEN_FILL, count=50)
place_card_back(sheet=workbook.create_sheet(title='F_back'), color=MED_BLUE_FILL, count=8)
place_card_back(sheet=workbook.create_sheet(title='O_back'), color=MED_YELLOW_FILL, count=16)
place_card_back(sheet=workbook.create_sheet(title='E_back'), color=MED_RED_FILL, count=14)

troubles = Troubles()
troubles.generate_troubles(count=28, rate=0)
troubles.generate_troubles(count=14, rate=1)
troubles.generate_troubles(count=14, rate=2)
troubles.generate_troubles(count=14, rate=3)

for troubles_list in troubles.troubles_list:
    sheet = workbook.create_sheet(troubles_list[0].letter_index)
    place_troubles_on_sheet(troubles_list=troubles_list, sheet=sheet)

place_card_back(sheet=workbook.create_sheet(title='T_back'), color=MED_BLACK_FILL, count=70)


modificators = Modificators()
modificators.generate_modificators(count=30)
for modificators_list in modificators.modificators_list:
    sheet = workbook.create_sheet(modificators_list[0].letter_index)
    place_modoficators_on_sheet(mod_list=modificators_list, sheet=sheet)
place_card_back(sheet=workbook.create_sheet(title='M_back'), color=MED_BIRUZA, count=30)
workbook.save("output.xlsx")
