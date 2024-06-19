from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from Styles import THICK_BORDER, THIN_BORDER, MED_GREEN_FILL, MED_BLUE_FILL, MED_YELLOW_FILL, MED_RED_FILL, BLUE_FILL, \
    GREEN_FILL, RED_FILL, MED_BLACK_FILL, NO_FILL
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from optimization_catalog import gen


def set_dim_holder(ws, start=1, end=50, width=2 * 1.35):
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(start, end):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, width=width, index=get_column_letter(col))

    return dim_holder


def place_story_on_sheet(stories_list, sheet, columns=[1, 19]):
    sheet.column_dimensions = set_dim_holder(ws=sheet)

    row = -9
    for i, story in enumerate(stories_list):
        # распределяем сторисы по двум колонкам
        if not i % 2:
            start_column = columns[0]
            row += 10
        else:
            start_column = columns[1]

        # Write outline border and fill the card with color
        write_outline_border(sheet=sheet, start_row=row, start_column=start_column, size={'c': 17, 'r': 9})
        match story.letter_index:
            case 'S':
                fill = MED_GREEN_FILL
            case 'F':
                fill = MED_BLUE_FILL
            case 'O':
                fill = MED_YELLOW_FILL
            case 'E':
                fill = MED_RED_FILL
        color_cells(sheet=sheet, start_row=row, start_column=start_column, size={'c': 17, 'r': 9}, fill=fill)

        # Write cells for analyst, developer and tester estimates and fill it with color
        write_all_border(sheet=sheet, start_row=row + 3, start_column=start_column + 1,
                         size={'c': story.analyst_estimate, 'r': 1})
        color_cells(sheet=sheet, start_row=row + 3, start_column=start_column + 1, size={'c': story.analyst_estimate, 'r': 1}, fill=RED_FILL)
        write_all_border(sheet=sheet, start_row=row + 5, start_column=start_column + 1,
                         size={'c': story.dev_estimate, 'r': 1})
        color_cells(sheet=sheet, start_row=row + 5, start_column=start_column + 1,
                    size={'c': story.dev_estimate, 'r': 1}, fill=BLUE_FILL)
        write_all_border(sheet=sheet, start_row=row + 7, start_column=start_column + 1,
                         size={'c': story.test_estimate, 'r': 1})
        color_cells(sheet=sheet, start_row=row + 7, start_column=start_column + 1,
                    size={'c': story.test_estimate, 'r': 1}, fill=GREEN_FILL)

        story_num_cell = sheet.cell(row=row, column=start_column)
        story_num_cell.value = story.letter_index + str(story.num)
        story_num_cell.alignment = Alignment(horizontal='center')
        story_num_cell.font = Font(bold=True, size=12)
        sheet.merge_cells(start_column=start_column, end_column=start_column + 2, start_row=row, end_row=row)

        price_cell = sheet.cell(row=row, column=start_column + 13)
        price_cell.value = '$' + str(story.price) if story.price else str(story.stability_score) + ' о.с.'
        price_cell.alignment = Alignment(horizontal='center')
        price_cell.font = Font(bold=True, size=12)
        sheet.merge_cells(start_column=start_column + 13, end_column=start_column + 16, start_row=row, end_row=row)

        title_cell = sheet.cell(row=row + 1, column=start_column + 1)
        title = ''
        match story.letter_index:
            case 'S':
                title = 'Стандарная история'
            case 'F':
                title = f'Крайний срок: {story.due_day} день'
            case 'O':
                title = next(gen)
            case 'E':
                title = f'Срок: {story.due_day} дня'

        title_cell.value = title
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.font = Font(bold=True, size=12)
        sheet.merge_cells(start_column=start_column + 1, end_column=start_column + 15, start_row=row + 1,
                          end_row=row + 1)


def place_troubles_on_sheet(troubles_list, sheet, columns=[1, 19]):
    sheet.column_dimensions = set_dim_holder(ws=sheet)
    row = -9
    for i, trouble in enumerate(troubles_list):
        # распределяем карточки по двум колонкам
        if not i % 2:
            start_column = columns[0]
            row += 10
        else:
            start_column = columns[1]

        # Write outline border and fill the card with color
        write_outline_border(sheet=sheet, start_row=row, start_column=start_column, size={'c': 17, 'r': 9})

        # Place the text
        text_cell = sheet.cell(row=row + 1, column=start_column + 1)
        text = trouble.text
        text_cell.value = text
        text_cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        text_cell.font = Font(bold=False, size=10)
        sheet.merge_cells(start_column=start_column + 1, end_column=start_column + 15, start_row=row + 1,
                          end_row=row + 7)

#       Place the price
        price_cell = sheet.cell(row=row, column=start_column + 13)
        price_cell.value = str(trouble.price) + ' о.с.'
        price_cell.alignment = Alignment(horizontal='center')
        price_cell.font = Font(bold=True, size=12)
        sheet.merge_cells(start_column=start_column + 13, end_column=start_column + 16, start_row=row, end_row=row)

# Place the rate
        story_num_cell = sheet.cell(row=row, column=start_column)
        story_num_cell.value = trouble.letter_index + str(trouble.rate)
        story_num_cell.alignment = Alignment(horizontal='center')
        story_num_cell.font = Font(bold=True, size=12)
        sheet.merge_cells(start_column=start_column, end_column=start_column + 2, start_row=row, end_row=row)


def place_modoficators_on_sheet(mod_list, sheet, columns=[1, 19]):
    sheet.column_dimensions = set_dim_holder(ws=sheet)
    row = -9
    for i, mod in enumerate(mod_list):
        # распределяем карточки по двум колонкам
        if not i % 2:
            start_column = columns[0]
            row += 10
        else:
            start_column = columns[1]

        # Write outline border and fill the card with color
        write_outline_border(sheet=sheet, start_row=row, start_column=start_column, size={'c': 17, 'r': 9})

        # Place the text
        text_cell = sheet.cell(row=row + 1, column=start_column + 1)
        text = mod.text
        text_cell.value = text
        text_cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        text_cell.font = Font(bold=False, size=10)
        sheet.merge_cells(start_column=start_column + 1, end_column=start_column + 15, start_row=row + 1,
                          end_row=row + 7)

def write_outline_border(sheet, start_row, start_column, size):
    stop_row = start_row + size['r']
    stop_column = start_column + size['c']
    for row in range(start_row, stop_row):
        for column in range(start_column, stop_column):
            left_side = THICK_BORDER if column == start_column else None
            right_side = THICK_BORDER if column == stop_column - 1 else None
            top_side = THICK_BORDER if row == start_row else None
            bottom_side = THICK_BORDER if row == stop_row - 1 else None
            sheet.cell(row=row, column=column).border = Border(left=left_side, right=right_side, top=top_side,
                                                               bottom=bottom_side)


def write_all_border(sheet, start_row, start_column, size):
    stop_row = start_row + size['r']
    stop_column = start_column + size['c']
    for row in range(start_row, stop_row):
        for column in range(start_column, stop_column):
            left_side = THIN_BORDER
            right_side = THIN_BORDER
            top_side = THIN_BORDER
            bottom_side = THIN_BORDER
            sheet.cell(row=row, column=column).border = Border(left=left_side, right=right_side, top=top_side,
                                                               bottom=bottom_side)


def color_cells(sheet, start_row, start_column, size, fill):
    stop_row = start_row + size['r']
    stop_column = start_column + size['c']
    for row in range(start_row, stop_row):
        for column in range(start_column, stop_column):
            sheet.cell(row=row, column=column).fill = fill


def place_card_back(sheet=None, color=NO_FILL, text='', count=2, columns=[1, 19]):
    sheet.column_dimensions = set_dim_holder(ws=sheet)
    row = -9
    for i in range(0,count):
        # распределяем карточки по двум колонкам
        if not i % 2:
            start_column = columns[0]
            row += 10
        else:
            start_column = columns[1]

        # Write outline border and fill the card with color
        write_outline_border(sheet=sheet, start_row=row, start_column=start_column, size={'c': 17, 'r': 9})

        # Place the text
        text_cell = sheet.cell(row=row + 1, column=start_column + 1)
        text_cell.value = text
        text_cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        text_cell.font = Font(bold=False, size=50)
        sheet.merge_cells(start_column=start_column, end_column=start_column+16, start_row=row,
                          end_row=row + 8)

        # FILL
        color_cells(sheet=sheet, start_row=row, start_column=start_column, size={'c': 17, 'r': 9}, fill=color)